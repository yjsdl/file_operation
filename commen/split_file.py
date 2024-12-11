# -*- coding: utf-8 -*-
# @date：2023/5/25 15:26
# @Author：LiuYiJie
# @file： split_file
import os

from openpyxl import load_workbook, Workbook
from openpyxl.utils import range_boundaries
import pandas as pd
from logger_code.log import logger


class fileSplit:
    def __init__(self, dir_path:str =None, file_path:str=None, dest_path:str =None):
        self.dir_path = dir_path
        self.file_path = file_path
        self.file_name = None
        self.dest_path = dest_path

    def execl_split(self):
        # 定义每个文件的最大行数
        batch_size = 100000

        path = r'E:\任务数据\爱学术\图书信息.csv'
        # 打开Excel文件
        wb = load_workbook(path)

        # 获取第一个worksheet对象
        ws = wb.worksheets[0]

        # 获取打开文件的范围，即worksheet的行列数
        min_row, min_col, max_row, max_col = range_boundaries(ws.dimensions)

        # 计算需要分割的文件数量
        batch_count = (max_row - min_row) // batch_size + 1

        # 遍历每个文件并读取数据
        for batch_index in range(batch_count):
            # 计算当前文件的起始行和结束行
            start_row = min_row + batch_index * batch_size + 1
            end_row = min_row + (batch_index + 1) * batch_size

            # 创建新的Excel文件
            new_wb = Workbook()

            # 获取worksheet对象
            new_ws = new_wb.active

            # 遍历行并复制单元格数据到新的worksheet中
            for row in ws.iter_rows(min_row=start_row, min_col=min_col, max_row=end_row, max_col=max_col):
                row_data = [cell.value for cell in row]
                new_ws.append(row_data)

            # 保存新的Excel文件
            new_wb.save('path/to/new_excel_file_{}.xlsx'.format(batch_index + 1))


    def csv_split(self):
        # 定义每个小文件的行数
        chunk_size = 10000

        # 读取大文件
        reader = pd.read_csv(self.file_path, chunksize=chunk_size, dtype=str)

        if not os.path.exists(dest_path):
            os.makedirs(dest_path)
        # 拆分文件并保存
        for i, chunk in enumerate(reader):
            res_path = os.path.join(dest_path, self.file_name.replace('.csv', f'__{i + 1}.csv'))
            # chunk now pf
            print(f'拆分文件{res_path}')
            chunk.to_csv(res_path, index=False, encoding='utf-8')

    def main(self):
        if self.dir_path:
            for root, dirs, files in os.walk(self.dir_path):
                for file in files:
                    self.file_name = file
                    self.file_path = os.path.join(root, file)
                    print(self.file_path)
                    self.csv_split()
        else:
            self.csv_split()



if __name__ == '__main__':
    start_path = r'Y:\wos_article2subject\wos_data_from_xml\WOS2021'
    dest_path = start_path.rsplit('\\', maxsplit=1)[0] + '\split\\' + start_path.rsplit('\\', maxsplit=1)[1]
    c = fileSplit(dir_path=start_path, dest_path=dest_path)
    c.main()
